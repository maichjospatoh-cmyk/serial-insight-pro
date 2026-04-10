import pandas as pd
import sys
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

file1 = sys.argv[1]
file2 = sys.argv[2]

def extract_serial(text):
    if pd.isna(text):
        return None
    text = str(text)
    match = re.search(r'89254021\d+', text)
    return match.group(0) if match else None

def extract_plate(text):
    if pd.isna(text):
        return "UNKNOWN"
    text = str(text)
    match = re.search(r'\bKC[A-Z]\b', text)
    return match.group(0) if match else "UNKNOWN"

def read_file(file):
    try:
        return pd.read_excel(file)
    except:
        return pd.read_csv(file, encoding='latin1')

df1 = read_file(file1)
df2 = read_file(file2)

# 🔥 EXTRACT SERIAL + VAN PLATE
df1["serial"] = df1.astype(str).apply(lambda r: extract_serial(" ".join(r)), axis=1)
df2["serial"] = df2.astype(str).apply(lambda r: extract_serial(" ".join(r)), axis=1)

df1["van_plate"] = df1.astype(str).apply(lambda r: extract_plate(" ".join(r)), axis=1)
df2["van_plate"] = df2.astype(str).apply(lambda r: extract_plate(" ".join(r)), axis=1)

df1 = df1.dropna(subset=["serial"])
df2 = df2.dropna(subset=["serial"])

merged = pd.merge(df1, df2, on="serial", how="outer")

if "agent name" not in merged.columns:
    merged["agent name"] = "Unknown"

# 🔥 DUPLICATES
merged["duplicate_per_agent"] = merged.duplicated(subset=["agent name","serial"], keep=False)

output_file = "output/result.xlsx"
merged.to_excel(output_file, index=False)

# 🔥 FORMAT EXCEL
wb = load_workbook(output_file)
ws = wb.active

red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

headers = [c.value for c in ws[1]]
dup_idx = headers.index("duplicate_per_agent")

for row in ws.iter_rows(min_row=2):
    if row[dup_idx].value:
        for cell in row:
            cell.fill = red

# 🔥 DASHBOARD SHEET
summary = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    agent = row[headers.index("agent name")]
    plate = row[headers.index("van_plate_x")] if "van_plate_x" in headers else "UNKNOWN"
    dup = row[dup_idx]

    key = (agent, plate)

    if key not in summary:
        summary[key] = {"total":0,"dup":0}

    summary[key]["total"] += 1
    if dup:
        summary[key]["dup"] += 1

ws2 = wb.create_sheet("Dashboard")

ws2.append(["Agent","Van Plate","Total","Duplicates","Quality %"])

for (agent,plate),val in summary.items():
    total = val["total"]
    dup = val["dup"]
    quality = ((total-dup)/total*100) if total else 0
    ws2.append([agent,plate,total,dup,round(quality,1)])

wb.save(output_file)

print("DONE")
